import io, re, smtplib, ssl
from email.message import EmailMessage

import streamlit as st
from openpyxl import load_workbook

XLSX = "vnd.openxmlformats-officedocument.spreadsheetml.sheet"
st.set_page_config(page_title="Gmail 자동발송", page_icon="📧")


# ── 공통 헬퍼 ────────────────────────────────────────────────
def secret(key: str) -> str:
    """secrets.toml이 없어도 앱이 죽지 않게 안전 조회"""
    try:
        return st.secrets[key]
    except Exception:
        return ""


def parse_emails(text: str) -> list:
    """콤마·세미콜론·줄바꿈·공백 아무거나로 구분, 중복 제거"""
    found = [e.strip() for e in re.split(r"[,;\s]+", text or "") if "@" in e]
    return list(dict.fromkeys(found))


def smart(v: str):
    """엑셀에 숫자는 숫자로 넣기 (안 그러면 문자열로 들어가 합계가 안 됨)"""
    for cast in (int, float):
        try:
            return cast(v)
        except (ValueError, TypeError):
            continue
    return v


for k in ("to_box", "cc_box", "bcc_box"):
    st.session_state.setdefault(k, "")


# ── 사이드바 : 계정 + 파일 ───────────────────────────────────
with st.sidebar:
    st.header("🔐 발신 계정")
    user_in = st.text_input("Gmail 주소")
    pw_in = st.text_input("앱 비밀번호(16자리)", type="password")

    # 수기입력이 우선, 비어 있으면 secrets.toml 값 사용
    GMAIL_USER = user_in.strip() or secret("GMAIL_USER")
    GMAIL_PW = pw_in.replace(" ", "") or secret("GMAIL_APP_PW")
    src = "수기입력" if user_in.strip() else ("secrets" if GMAIL_USER else "미설정")
    st.caption(f"발신: **{GMAIL_USER or '—'}** ({src})")

    st.divider()
    st.header("📎 첨부할 엑셀")
    up = st.file_uploader("수정·첨부 파일(.xlsx)", type="xlsx", key="attach")
    sheet = None
    if up:
        up.seek(0)
        sheet = st.selectbox("시트 선택", load_workbook(up).sheetnames)

    st.divider()
    st.header("👥 수신자 엑셀 (A열)")
    rcpt_file = st.file_uploader("이메일 목록(.xlsx)", type="xlsx", key="rcpt")
    target = st.radio("넣을 위치", ["수신", "참조", "숨은참조"], horizontal=True)
    skip_head = st.checkbox("첫 행은 머리글", value=True)
    mode = st.radio("적재 방식", ["덮어쓰기", "이어붙이기"], horizontal=True)

    if st.button("목록 불러오기", disabled=not rcpt_file, width="stretch"):
        rcpt_file.seek(0)
        ws = load_workbook(rcpt_file, data_only=True).active
        emails = []
        for i, row in enumerate(ws.iter_rows(min_col=1, max_col=1, values_only=True)):
            if skip_head and i == 0:
                continue
            v = str(row[0] or "").strip()
            if "@" in v:
                emails.append(v)

        key = {"수신": "to_box", "참조": "cc_box", "숨은참조": "bcc_box"}[target]
        merged = (
            parse_emails(st.session_state[key]) + emails
            if mode == "이어붙이기"
            else emails
        )
        st.session_state[key] = ", ".join(dict.fromkeys(merged))
        st.success(f"{len(emails)}건 → {target}")


# ── 본문 : 수신자 ────────────────────────────────────────────
st.title("📧 엑셀 수정 → Gmail 자동발송")

st.subheader("1. 수신자")
to_txt = st.text_area("수신 (To)", key="to_box", height=68)
c1, c2 = st.columns(2)
cc_txt = c1.text_area("참조 (Cc)", key="cc_box", height=68)
bcc_txt = c2.text_area("숨은참조 (Bcc)", key="bcc_box", height=68)

to_list, cc_list, bcc_list = map(parse_emails, (to_txt, cc_txt, bcc_txt))
st.caption(f"수신 {len(to_list)} · 참조 {len(cc_list)} · 숨은참조 {len(bcc_list)}명")


# ── 본문 : 셀 수정 (3개 이상) ────────────────────────────────
st.subheader("2. 엑셀 수정 항목")
n = st.number_input("수정할 셀 개수", min_value=1, max_value=30, value=3, step=1)

DEFAULTS = [("B2", "2026-08-07"), ("C5", "1250000"), ("D7", "김대리")]
edits = []
for i in range(int(n)):
    dc, dv = DEFAULTS[i] if i < len(DEFAULTS) else ("", "")
    a, b = st.columns([1, 3])
    cell = a.text_input(f"셀 {i+1}", dc, key=f"cell{i}", placeholder="B2")
    val = b.text_input(f"값 {i+1}", dv, key=f"val{i}")
    if cell.strip():
        edits.append((cell.strip().upper(), val))


# ── 본문 : 메일 내용 ─────────────────────────────────────────
st.subheader("3. 메일 내용")
subject = st.text_input("제목", "[보고] 월간 실적")
body = st.text_area("본문", "안녕하세요.\n첨부 파일 확인 부탁드립니다.")

ready = all([up, GMAIL_USER, GMAIL_PW, to_list or cc_list or bcc_list])
if not ready:
    st.info("사이드바에서 계정·첨부파일을 설정하고 수신자를 1명 이상 입력하세요.")

if st.button("✉️ 수정 후 발송", type="primary", disabled=not ready, width="stretch"):
    try:
        # 1) 메모리에서 엑셀 수정
        up.seek(0)
        wb = load_workbook(up)
        ws = wb[sheet]
        for cell, val in edits:
            ws[cell] = smart(val)
        buf = io.BytesIO()
        wb.save(buf)

        # 2) 메일 작성 (Bcc 헤더는 send_message가 발송 시 자동 제거)
        msg = EmailMessage()
        msg["From"] = GMAIL_USER
        msg["Subject"] = subject
        if to_list:
            msg["To"] = ", ".join(to_list)
        if cc_list:
            msg["Cc"] = ", ".join(cc_list)
        if bcc_list:
            msg["Bcc"] = ", ".join(bcc_list)
        msg.set_content(body)
        msg.add_attachment(
            buf.getvalue(), maintype="application", subtype=XLSX, filename=up.name
        )

        # 3) 발송
        with smtplib.SMTP_SSL(
            "smtp.gmail.com", 465, context=ssl.create_default_context(), timeout=30
        ) as s:
            s.login(GMAIL_USER, GMAIL_PW)
            s.send_message(msg)

        total = len(to_list) + len(cc_list) + len(bcc_list)
        st.success(f"✅ 총 {total}명에게 발송 완료 (수정 셀 {len(edits)}개)")
        st.download_button("수정본 내려받기", buf.getvalue(), file_name=up.name)

    except KeyError as e:
        st.error(f"❌ 셀/시트 주소 오류: {e}")
    except smtplib.SMTPAuthenticationError:
        st.error("❌ 인증 실패 — 앱 비밀번호 16자리가 맞는지, 2단계 인증이 켜져 있는지 확인하세요.")
    except Exception as e:
        st.error(f"❌ 실패: {e}")
